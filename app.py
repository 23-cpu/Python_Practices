import openpyxl as xl
from openpyxl.chart import BarChart, Reference


work_book = xl.load_workbook("transactions.xlsx")
sheet_1 = work_book['Sheet1']
cell = sheet_1['a2']

# Another way of accessing the cell
# cell_1 = sheet_1.cell(1,1)
# print(cell_1.value)

# correcting a faulty group of cells in the worksheet
for row in range(2, sheet_1.max_row+1):
    cell = sheet_1.cell(row, 3)
    corrected_values = cell.value * 0.9

    corrected_price_cell = sheet_1.cell(row, 4)
    corrected_price_cell.value = corrected_values

#Getting the values in the fourth column
Values = Reference(sheet_1,
          min_row=2,
          max_row=sheet_1.max_row,
          min_col=4,
          max_col=4)


chart = BarChart()
chart.add_data(Values)
sheet_1.add_chart(chart, 'e2')

work_book.save('transaction2data.xlsx')
   
    
 